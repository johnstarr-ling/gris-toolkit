from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import argparse 
import numpy as np 
import pandas as pd 
import math 



###############################################################################
# Functions
###############################################################################


##### PROCESSING EXCEL SHEET
def load_sheet(filename, sheetname):
    
    # Load workbook
    workbook = load_workbook(filename)
    
    return workbook[sheetname] # Return the sheet that we want


def make_color_array(sheet, num_rows, num_cols):

    # Build the canvas as an np array
    array = np.empty((num_rows, num_cols), dtype=object)
    
    # Assign the color labels to the position
    for row in range(1, num_rows+1):
        for column in range(1, num_cols+1):
            color = sheet[get_column_letter(column)+str(row)].fill.start_color.index
            array[row-1, column-1] = f'#{color[2:]}'
    
    return array


def make_location_array(x_positions, y_positions):

    # Build the canvas as an np array
    array = np.empty((len(y_positions), len(x_positions)), dtype=object)
    
    # Assign the location labels to the position
    for row in range(len(x_positions)):
        for column in range(len(y_positions)):
            array[column, row] = (y_positions[column], x_positions[row])
    
    return array


def get_cluster_bounds(arr):
    bounds = {}
    for val in np.unique(arr):
        positions = np.argwhere(arr == val)
        r1, c1 = positions.min(axis=0)
        r2, c2 = positions.max(axis=0) + 1  # exclusive
        bounds[val] = (r1, r2, c1, c2)
    return bounds


def assign_cluster_indices(bounds):
    sorted_vals = sorted(bounds.items(), key=lambda x: (x[1][0], x[1][2]))
    cluster_rows = []
    coords = {}

    for val, (r1, r2, c1, c2) in sorted_vals:
        # Find a row with the same top row (r1)
        for row_idx, row in enumerate(cluster_rows):
            if bounds[row[0]][0] == r1:
                col_idx = len(row)
                row.append(val)
                coords[val] = (row_idx, col_idx)
                break
        else:
            # New cluster row
            cluster_rows.append([val])
            coords[val] = (len(cluster_rows) - 1, 0)

    return coords


# Make label arrays
def make_label_array(arr):

    arr = np.array(arr)
    bounds = get_cluster_bounds(arr)
    cluster_coords = assign_cluster_indices(bounds)

    output = np.empty(arr.shape, dtype=object)
    for val, (r1, r2, c1, c2) in bounds.items():
        x_cluster, y_cluster = cluster_coords[val]
        for i in range(r1, r2):
            for j in range(c1, c2):
                output[i, j] = (x_cluster, y_cluster, i, j)
    
    return output


##### GET LOCATIONS OF CELLS
def generate_range(midpoint, step, num_X):
    
    # num_X is the number of rows/columns that we want.
    # Repetitions is the number of canvases we want on each side
    # of the midpoint.
    
    step = int(step)
    repetitions = num_X/2 

    # If repetitions isn't an integer, set the first cell to cross the midpoint.
    if repetitions != int(repetitions):
        midpoint = midpoint - (step / 2)
    
    # Calculate steps from the midpoint 
    int_repetitions = math.floor(repetitions)
    left_steps = [midpoint - i * step for i in range(int_repetitions, -1, -1)]
    
    # Odd number of repetitions have extra value (center)
    if repetitions != int(repetitions):
        right_steps = [midpoint + i * step for i in range(1, int_repetitions + 1)]
    
    # Even number of repetitions have the same number of X on other side of the midpoint
    else:
        right_steps = [midpoint + i * step for i in range(1, int_repetitions)]
    
    return left_steps + right_steps



# Centers canvas both horizontally/vertically; x-range (250, 1270) | y-range (50, 470)
def canvas_positions(num_rows, num_cols,
                     col_width, row_height,
                     x_min, x_max, 
                     y_min, y_max):
    
    # Spread outward from the center:
    x_mid = (x_max+x_min)/2 
    y_mid = (y_max+y_min)/2
    
    # Column positions 
    col_positions = generate_range(x_mid, col_width, num_cols) 
    
    # Row positions    
    row_positions = generate_range(y_mid, row_height, num_rows)

    return col_positions, row_positions 



##### WRITING IBEX CODE 
# Constructs newCanvas objects
def make_newCanvas(label, xlocation, ylocation, color, 
                   width, height, autofit=False):

    if autofit == True:
        return f'newCanvas("{label}", "{width}vw", "{height}vh").color("{color}").print("{xlocation}", "{ylocation}"),'
    
    return f'newCanvas("{label}", {width}, {height}).color("{color}").print({xlocation}, {ylocation}),'

# Constructs getCanvas objects
def make_getCanvas(label):
    return f'getCanvas("{label}"),'


##### MAIN 
if __name__ == '__main__':
    
    # Argparse 
    parser = argparse.ArgumentParser()

    parser.add_argument('-f', '--file', type=str,
                        default='./data/gris-demos.xlsx',
                        help='Input xlsx file')
    
    parser.add_argument('-s', '--sheet', type=str,
                        default='blank10x10',
                        help='Sheet name from input xlsx file')
    
    parser.add_argument('-o', '--outputname', type=str,
                        default='./outputs/xlsx_output.txt', 
                        help='Output filename.')
    
    parser.add_argument('-cw', '--width', type=str,
                        default=30,
                        help='Column width (in pixels.')

    parser.add_argument('-hr', '--height', type=str, 
                        default=30,
                        help='Row height (in pixels).')

    parser.add_argument('-xmin', '--x_minimum', type=int,
                        default=250,
                        help='Leftmost bound of the canvas on the screen.')

    parser.add_argument('-xmax', '--x_maximum', type=int,
                        default=1270,
                        help='Rightmost bound of canvas on the screen.')

    parser.add_argument('-ymin', '--y_minimum', type=int,
                        default=50,
                        help='Topmost bound of the canvas on the screen.')

    parser.add_argument('-ymax', '--y_maximum', type=int,
                        default=470,
                        help='Bottommost bound of the canvas on the screen.')

    parser.add_argument('-a', '--auto_fit', action='store_true',
                        default=False,
                        help='Build that automatically scales.')

    args = parser.parse_args()

    FILE = args.file
    SHEET = args.sheet
    OUTPUT = args.outputname
    WIDTH = int(args.width)
    HEIGHT = int(args.height)
    XMIN = args.x_minimum 
    XMAX = args.x_maximum
    YMIN = args.y_minimum
    YMAX = args.y_maximum

    if args.auto_fit:
        XMIN = 0
        XMAX = 100
        YMIN = 0 
        YMAX = 100


    # Processing sheet
    sheet = load_sheet(FILE, SHEET)
    
    NUM_ROWS = sheet.max_row 
    NUM_COLS = sheet.max_column 

    if args.auto_fit:
        # Ensure autofit boundaries to reduce scrolling
        if NUM_COLS*WIDTH >= 100:
            raise Exception(f'Width specification too wide for autofit specification; please ensure the total width <100vh. Currently: {NUM_COLS*WIDTH}')
        if NUM_ROWS*HEIGHT >= 100:
            raise Exception(f'Height specification too wide for autofit specification; please ensure the total height <100 vh. Currently: {NUM_ROWS*HEIGHT}')

    # Make color array
    color_array = make_color_array(sheet, NUM_ROWS, NUM_COLS)
    
    # Make label array:
    label_array = make_label_array(color_array)

    # Drawing grid 
    x_positions, y_positions = canvas_positions(NUM_ROWS, NUM_COLS,
                                                WIDTH, HEIGHT,
                                                XMIN, XMAX,
                                                YMIN, YMAX)

    # Build location array
    location_array = make_location_array(x_positions, y_positions)

    # Adding newCanvas Header
    canvas_string = '############### NEW CANVAS ###############\n'

    # newCanvas loop
    for i in range(NUM_ROWS):
        for j in range(NUM_COLS):
            if args.auto_fit:
                canvas_string += make_newCanvas(label_array[i, j], f'{location_array[i, j][1]}vw', f'{location_array[i, j][0]}vh', color_array[i, j], WIDTH, HEIGHT, autofit=True)
            else:
                canvas_string += make_newCanvas(label_array[i, j], location_array[i, j][1], location_array[i, j][0], color_array[i, j], WIDTH, HEIGHT)

        canvas_string += '\n'

    # Adding getCanvas Header
    canvas_string += '\n\n\n\n\n############### GET CANVAS ###############\n'

    # getCanvas loop
    for i in range(NUM_ROWS):
        for j in range(NUM_COLS):
            canvas_string += make_getCanvas(label_array[i, j])
        canvas_string += '\n'

    # Writing to output file
    with open(OUTPUT, 'w') as w:
        w.write(canvas_string)